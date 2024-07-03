"""Tests for the extract module."""

import csv
import os
from pathlib import Path

import openpyxl
import pytest

from heave import file, utils
from heave.file import Table


class TestTable:
    """Test the Table class."""

    def test_init(self):
        """Test initialisation of the Table class."""
        data = [("header1", "header2"), ("data1", "data2")]
        table = Table(data)
        assert table.header == ("header1", "header2")
        assert list(table.rows) == [("data1", "data2")]


class TestCsv:
    """Test the read_csv function."""

    test_file = Path("temp.csv")

    @pytest.fixture(autouse=True)
    def temp_file(self):
        """Create a temporary csv file."""
        with open(self.test_file, "w") as f:
            f.write("header1,header2\n")
            f.write("data1,data2\n")
        yield
        os.remove(self.test_file)

    def test_read_csv(self):
        """Test reading a csv file."""
        table = file.read_csv(self.test_file)
        assert table.header == ("header1", "header2")
        assert list(table.rows) == [("data1", "data2")]

    def test_write_csv(self):
        """Test writing a csv file."""
        data = Table([("header3", "header4"), ("data3", "data4")])
        file.write_csv(data, self.test_file)
        table = file.read_csv(self.test_file)
        assert table == data


class TestXlsx:
    """Test xlsx to csv conversion."""

    test_file = Path("temp1.xlsx")

    @pytest.fixture(autouse=True)
    def temp_file(self):
        """Create a temporary xlsx file and yield its directory."""
        workbook = openpyxl.Workbook()

        # First sheet
        worksheet1 = workbook.create_sheet("1")
        worksheet1["A1"] = "header1"
        worksheet1["B1"] = "header2"
        worksheet1["A2"] = "data1"
        worksheet1["B2"] = "data2"
        # Second sheet
        worksheet2 = workbook.create_sheet("2")
        worksheet2["A1"] = "heading1"
        worksheet2["B1"] = "heading2"
        worksheet2["C1"] = "heading3"
        worksheet2["A2"] = "data1"
        worksheet2["B2"] = "data2"
        worksheet2["C2"] = "data3"

        workbook.save(self.test_file)

        yield
        os.remove(self.test_file)

    def test_xlsx_to_csv(self):
        """Test converting xlsx to csv using xlsx_to_csv function."""
        # First sheet and its content
        csv_file1 = utils.xlsx_to_csv(self.test_file, sheet_name="1")
        assert csv_file1.exists()
        with open(csv_file1) as f:
            reader = csv.reader(f)
            data = list(reader)
        assert len(data[0]) == 2  # Check number of columns
        assert data[0] == ["header1", "header2"]
        assert data[1] == ["data1", "data2"]

        # Second sheet and its content
        csv_file2 = utils.xlsx_to_csv(self.test_file, sheet_name="2")
        assert csv_file2.exists()
        with open(csv_file2) as f:
            reader = csv.reader(f)
            data = list(reader)
        assert len(data[0]) == 3  # Check number of columns
        assert data[0] == ["heading1", "heading2", "heading3"]
        assert data[1] == ["data1", "data2", "data3"]

        os.remove(csv_file1)
        os.remove(csv_file2)
